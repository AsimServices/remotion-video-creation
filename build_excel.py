#!/usr/bin/env python3
"""
build_excel.py
Generates a beautifully formatted Excel workbook with all Remotion video
categories and ideas — both existing (from video_ideas.txt) and new creative ones.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.numbers import FORMAT_TEXT

# ── Palette ──────────────────────────────────────────────────────────────────
BG_DARK        = "0D0D0D"
BG_SHEET       = "111827"
ACCENT_CYAN    = "00E5FF"
ACCENT_GOLD    = "FFD700"
ACCENT_PURPLE  = "9B59B6"
ACCENT_GREEN   = "2ECC71"
ACCENT_ORANGE  = "FF6B35"
ACCENT_PINK    = "FF4081"
ACCENT_BLUE    = "2196F3"
ACCENT_TEAL    = "00BCD4"
ACCENT_LIME    = "CDDC39"
ACCENT_RED     = "F44336"
ACCENT_INDIGO  = "5C6BC0"
ACCENT_AMBER   = "FFC107"
WHITE          = "FFFFFF"
LIGHT_GRAY     = "E8EAF6"
MID_GRAY       = "9E9E9E"
DARK_GRAY      = "1E2A38"
ROW_ALT_A      = "1A2535"
ROW_ALT_B      = "141E2E"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border_thin(color="2C3E50"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium(color=ACCENT_CYAN):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def apply_header_row(ws, row_num, values, fills, font_colors=None, heights=None):
    for col, (val, bg) in enumerate(zip(values, fills), 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill = fill(bg)
        fc = font_colors[col-1] if font_colors else WHITE
        cell.font = font(bold=True, color=fc, size=12)
        cell.alignment = center()
        cell.border = border_thin(ACCENT_CYAN)
    if heights:
        ws.row_dimensions[row_num].height = heights

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

# ═══════════════════════════════════════════════════════════════════════════════
# DATA
# ═══════════════════════════════════════════════════════════════════════════════

CATEGORY_COLORS = {
    "Global Network Animations":          ACCENT_CYAN,
    "GPS / Route Tracing":                ACCENT_GREEN,
    "Data Point Pulses":                  ACCENT_GOLD,
    "Globe with Data Overlays":           ACCENT_BLUE,
    "Infographic Overlays":               ACCENT_ORANGE,
    "Flight Paths":                       ACCENT_PINK,
    "Heatmap Flows":                      ACCENT_RED,
    "Connection Webs":                    ACCENT_PURPLE,
    "Timeline Evolutions":                ACCENT_TEAL,
    "Icon Swarms":                        ACCENT_LIME,
    "Generative Mathematics":             ACCENT_INDIGO,
    "Particle Systems":                   ACCENT_CYAN,
    "Geometric / Abstract":               ACCENT_GOLD,
    "Physics Simulations":                ACCENT_GREEN,
    "Cyber / Tech Aesthetics":            ACCENT_BLUE,
    "Space & Cosmos":                     ACCENT_PURPLE,
    "Organic / Biological":               ACCENT_GREEN,
    "Architectural / Structural":         ACCENT_AMBER,
    "Audio-Visual Style":                 ACCENT_PINK,
    "Cinematic Transitions":              ACCENT_ORANGE,
    "Financial / Trading":                ACCENT_GOLD,
    "Sports & Events":                    ACCENT_RED,
}

EXISTING_DATA = {
    "Global Network Animations": {
        "description": "City & continent connections, business expansion, supply chains",
        "batches": [
            ("Ultra dark world map with razor-thin white connection lines and microscopic glowing city nodes", "Map/Geography"),
            ("Chrome metallic globe with mirror-finish arcs connecting gleaming city node spheres in space", "Globe/3D"),
            ("Deep navy world network with gold animated pulse routes and soft glowing halos on city nodes", "Map/Geography"),
            ("Aerial cinematic world map pull-back revealing full global network of glowing cyan connections", "Cinematic"),
        ]
    },
    "GPS / Route Tracing": {
        "description": "Moving dots, vehicles, logistics and shipping route animations",
        "batches": [
            ("Glowing neon GPS route line drawing itself across dark city grid with animated dot moving along path", "Navigation"),
            ("Electric blue navigation path tracing through dark road network with pulsing position marker", "Navigation"),
            ("Glowing delivery truck icon moving along neon route across dark city map with trail effect", "Logistics"),
            ("Container ship icon glowing in cyan tracing ocean shipping lane from port to port on dark map", "Maritime"),
            ("Multiple colored vehicle dots moving simultaneously along separate glowing routes on dark map", "Fleet"),
            ("Raw material to factory to store journey traced as glowing animated path across world map", "Supply Chain"),
            ("Glowing airplane icon tracing great circle arc route across dark globe with neon contrail", "Aviation"),
            ("Neon city grid with glowing taxi and rideshare vehicles tracing routes across dark urban map", "Urban Mobility"),
            ("Animated port operations map with glowing container ships queuing and docking in sequence", "Maritime"),
            ("Order fulfillment animation with glowing parcel dot traveling from warehouse to front door", "E-Commerce"),
            ("Cinematic drone-view GPS trace glowing in gold across dark photorealistic city grid at night", "Cinematic"),
        ]
    },
    "Data Point Pulses": {
        "description": "Markers growing or flashing on maps — sales, market share, performance",
        "batches": [
            ("Glowing cyan location pins dropping onto dark world map one by one with ripple pulse rings", "Map Markers"),
            ("Neon market presence dots expanding outward from city centers across dark regional map", "Market Share"),
            ("Electric gold sales markers flashing and growing on dark USA map showing revenue hotspots", "Revenue"),
            ("Glowing sales territory boundaries expanding outward from origin city on dark country map", "Territory"),
            ("Bright revenue hotspot markers exploding outward with particle bursts on dark city map", "Revenue"),
            ("Live sales data animation with markers randomly flashing across world map in real time style", "Real-Time"),
        ]
    },
    "Globe with Data Overlays": {
        "description": "Spinning Earth, satellite rings, financial flows orbiting a globe",
        "batches": [
            ("Spinning Earth with real-time style data arcs shooting between cities on dark background", "Globe"),
            ("Corporate globe with animated financial flow lines orbiting surface between finance hubs", "Finance"),
            ("Dark rotating Earth with satellite orbital rings circling and pulsing around globe equator", "Space"),
            ("Globe rotation with animated trade volume heat rings glowing brighter over busiest regions", "Trade"),
            ("Particle globe where continents are formed by thousands of tiny glowing dots on dark background", "Abstract"),
            ("Geometric low-poly Earth sphere rotating with flat-shaded continent faces glowing in gold", "Abstract"),
            ("Neon grid globe with only latitude and longitude lines glowing on perfectly black background", "Minimal"),
            ("Earth globe rising slowly from darkness below with city lights glowing on night side", "Cinematic"),
            ("Cinematic pull-back from city level rising through atmosphere to reveal full spinning Earth", "Cinematic"),
        ]
    },
    "Infographic Overlays": {
        "description": "Charts, bars, KPIs rising on maps — financial reports, demographics, stats",
        "batches": [
            ("Neon blue 3D bars rising from city locations on dark world map showing GDP by country", "Bar Charts"),
            ("Gold vertical bar columns growing from US state positions on dark map showing revenue data", "Revenue"),
            ("Dark world map with neon green stock market performance bars rising from financial capitals", "Finance"),
            ("Animated GDP growth bars rising from country centers with brightness scaling to performance", "Economics"),
            ("Animated regional revenue bars growing from sales territory centers on dark national map", "Sales"),
            ("Animated pie chart slices expanding outward from country centers on dark world map", "Pie Charts"),
            ("Glowing population density bars rising from city centers scaled to inhabitants on dark world map", "Demographics"),
            ("Dark world map with animated inflation rate bars rising from national capitals in red and gold", "Economics"),
            ("Animated KPI bar overlay with performance columns rising from office locations on dark map", "KPI"),
            ("Neon stock market index bars pulsing from exchange city locations on dark world map", "Finance"),
            ("Glowing number counters ticking upward above city markers on dark world map simultaneously", "Counters"),
            ("Cinematic camera push into dark world map as bar charts rise dramatically from all continents", "Cinematic"),
        ]
    },
    "Flight Paths": {
        "description": "Airplane icons following routes — travel, business aviation, air cargo",
        "batches": [
            ("Glowing airplane icon tracing neon arc route from New York to London on dark world map", "Single Route"),
            ("Gold airplane icon leaving glowing contrail arc as it crosses dark Atlantic on world map", "Transatlantic"),
            ("Major airline hub animation with glowing planes radiating outward from central airport node", "Hub & Spoke"),
            ("Ultra long-haul flight path animation with slow glowing aircraft crossing dark world map", "Long-Haul"),
            ("Corporate jet route animation with gold airplane icon tracing executive travel path on dark map", "Business Aviation"),
            ("Cargo aircraft animation with neon orange freighter icon tracing overnight route on dark map", "Air Cargo"),
            ("Summer holiday flight paths animation with planes streaming from Europe to Mediterranean", "Tourism"),
            ("Silk Road air trade animation with cargo planes following ancient routes on dark map", "Trade"),
            ("Rush hour air traffic animation with dense streams of neon aircraft crossing dark world map", "Real-Time Traffic"),
            ("European short-haul network animation with neon planes connecting capital cities on dark map", "Regional"),
            ("Cinematic globe rotation with glowing flight arc drawing slowly across dark Earth surface", "Cinematic"),
        ]
    },
    "Heatmap Flows": {
        "description": "Color gradients spreading across regions — density, risk, market analysis",
        "batches": [
            ("Red to yellow heat gradient slowly blooming outward from dense population centers on dark world map", "Population"),
            ("Cool blue to hot orange heatmap flow spreading across dark continental map from activity origins", "Market Density"),
            ("Risk exposure heatmap animation with red intensity spreading from volatile markets on dark world map", "Financial Risk"),
            ("Revenue density heatmap with gold intensity spreading from top-performing sales zones on dark map", "Revenue"),
            ("World population density heatmap animation with warm glow intensifying over megacity clusters", "Demographics"),
            ("Temperature anomaly heatmap animation with red intensity spreading over warming regions", "Climate"),
            ("Disease outbreak heatmap animation with red spread flowing outward from origin points", "Healthcare"),
            ("Internet usage density heatmap with cyan glow intensifying over connected regions on dark world map", "Digital"),
            ("Property value heatmap animation with gold intensity spreading from premium districts on dark city map", "Real Estate"),
            ("Dramatic heatmap ignition animation where dark world map slowly catches fire with intensity glow", "Cinematic"),
        ]
    },
    "Connection Webs": {
        "description": "Animated hubs and nodes for partnerships, networking, mergers & acquisitions",
        "batches": [
            ("Glowing central hub node with electric blue connection lines radiating outward to partner nodes", "Hub & Spoke"),
            ("Neon cyan network web with pulsing nodes and animated connection lines forming on pure black", "Network"),
            ("Two separate node networks slowly drifting together and merging into unified web on dark background", "M&A"),
            ("LinkedIn-style professional network web with glowing profile nodes and connection lines", "Social"),
            ("API integration web animation with platform hub node and app partner nodes connecting in sequence", "Tech Ecosystem"),
            ("Venture capital portfolio network with investor hub connected to startup nodes on dark background", "Finance"),
            ("Supplier network web animation with manufacturer hub connected to tier one and two suppliers", "Supply Chain"),
            ("Crystal node network animation with faceted gem-like hubs and prismatic connection lines", "Premium Abstract"),
            ("Corporate org chart animation with executive hub nodes and department sub-nodes connecting", "Org Structure"),
            ("Dramatic network convergence animation where scattered nodes fly together forming web", "Cinematic"),
        ]
    },
    "Timeline Evolutions": {
        "description": "Maps and networks changing over time — historical business & market growth",
        "batches": [
            ("Company founding to global expansion timeline animation with city markers appearing year by year", "Corporate History"),
            ("Ancient to modern trade route evolution animation with routes transforming on dark world map", "Trade History"),
            ("Internet adoption timeline animation with connectivity spreading across dark world map year by year", "Tech Adoption"),
            ("Retail chain expansion timeline with store location dots populating dark national map year by year", "Market Expansion"),
            ("Economic bloc formation timeline animation with country groups lighting up on dark world map", "Geopolitics"),
            ("Railway network construction timeline animation with track lines appearing on dark map over time", "Infrastructure"),
            ("City population growth timeline animation with urban zones expanding on dark regional map", "Demographics"),
            ("Fashion industry hub evolution timeline showing design capital importance shifting on dark world map", "Industry"),
            ("Deforestation timeline animation showing forest coverage shrinking on dark regional map over decades", "Environment"),
            ("Epic timeline sweep animation showing full century of business evolution on dark world map", "Cinematic"),
        ]
    },
    "Icon Swarms": {
        "description": "Icons clustering and moving by data — demographic and customer mapping",
        "batches": [
            ("Thousands of tiny glowing person icons clustering into dense population zones on dark world map", "Population"),
            ("Glowing customer icons flowing toward retail hub locations and clustering on dark city map", "Customer Behavior"),
            ("Age demographic icon swarm animation with icons color-coded by generation grouping on dark map", "Demographics"),
            ("Worker icons flowing from residential zones to employment hub clusters on dark city grid map", "Workforce"),
            ("Voter preference icon swarm animation with color-coded figures clustering into political zones", "Political"),
            ("Social media follower icons swarming toward influencer hub nodes on dark background", "Digital"),
            ("Patient icon swarms flowing toward hospital and clinic nodes across dark regional map", "Healthcare"),
            ("Survey respondent icons self-organizing into opinion clusters on dark background by response data", "Market Research"),
            ("Luxury gold person icons flowing in elegant streams toward premium destination nodes", "Premium"),
            ("Epic icon swarm convergence animation where thousands of figures fly together into globe shape", "Cinematic"),
        ]
    },
}

NEW_DATA = {
    "Generative Mathematics": {
        "description": "Fractal patterns, wave interference, parametric curves, and number theory visualizations",
        "color": ACCENT_INDIGO,
        "ideas": [
            ("Fractal animations — Sierpinski triangles, recursive branching trees, Koch snowflakes growing frame by frame", "Fractals"),
            ("Wave interference patterns — sine/cosine waves colliding, moiré patterns, standing waves", "Wave Math"),
            ("Lissajous figures — parametric curves drawing themselves in glowing neon", "Parametric"),
            ("Spirograph / harmonograph — rotating gear paths tracing intricate patterns", "Spirograph"),
            ("Perlin noise fields — flowing particle fields driven by deterministic noise grids", "Noise Fields"),
            ("Strange attractors — Lorenz, Rössler, Clifford attractors plotting in 3D-projected space", "Chaos Theory"),
            ("Prime number spirals — Ulam spiral, Sacks spiral, number-theoretic patterns", "Number Theory"),
            ("Golden ratio / Fibonacci — spiral growth, sunflower packing, nautilus animations", "Fibonacci"),
            ("Voronoi diagrams — cells forming, shifting, shattering on dark background", "Geometry"),
            ("Delaunay triangulation — mesh building and morphing animations", "Geometry"),
        ]
    },
    "Particle Systems": {
        "description": "Galaxy simulations, nebulae, fireworks, vortexes, and fluid-like particle animations",
        "color": ACCENT_CYAN,
        "ideas": [
            ("Galaxy simulations — star clusters orbiting dark centers with trail effects", "Cosmos"),
            ("Nebula formations — particle clouds expanding and contracting", "Cosmos"),
            ("Fireworks — deterministic burst patterns, color cascades", "Celebration"),
            ("Vortex / tornado — particles spiraling inward or outward", "Physics"),
            ("Magnetic field lines — dipole fields with animated particle tracers", "Physics"),
            ("Flock / swarm behavior — boid-style group motion (fully deterministic)", "Behavior"),
            ("Fluid simulation — particle-based water, smoke, or ink diffusion effects", "Fluid"),
            ("Explosion shockwaves — radial particle bursts with ring propagation", "Effects"),
            ("Black hole accretion — particles spiraling toward a dark center with lensing effect", "Cosmos"),
            ("Aurora borealis — wave curtains of glowing color on pure black sky", "Nature"),
        ]
    },
    "Geometric / Abstract": {
        "description": "Morphing polyhedra, kaleidoscopes, tessellations, and op-art illusion animations",
        "color": ACCENT_GOLD,
        "ideas": [
            ("Morphing polyhedra — cube → dodecahedron → sphere transitions", "3D Geometry"),
            ("Kaleidoscope — radially symmetrical geometric patterns rotating and evolving", "Symmetry"),
            ("Tesseract / 4D rotation — hypercube projected into 2D rotating in 4D", "4D Math"),
            ("Celtic knots — interlocking path animations drawing themselves", "Patterns"),
            ("Penrose tiling — aperiodic tiling patterns building across the frame", "Tiling"),
            ("Islamic geometric patterns — star polygon networks expanding", "Cultural"),
            ("Op-art animations — Victor Vasarely-style illusion grids pulsing and warping", "Illusions"),
            ("Concentric shape pulses — rings/squares/hexagons radiating outward", "Pulses"),
            ("Grid distortion fields — Cartesian grids warping, bending, folding", "Distortion"),
            ("Origami unfolding — geometric fold sequences in 2D projection", "Origami"),
        ]
    },
    "Physics Simulations": {
        "description": "Pendulum waves, orbital mechanics, elastic collisions, cellular automata",
        "color": ACCENT_GREEN,
        "ideas": [
            ("Pendulum waves — multiple pendulums of different lengths creating wave patterns", "Mechanics"),
            ("Double pendulum chaos — deterministic chaotic path tracing", "Chaos"),
            ("Orbital mechanics — planets orbiting with Kepler trajectories", "Orbital"),
            ("Elastic collision simulation — billiard-ball style particle bouncing", "Collisions"),
            ("Spring-mass systems — networks of nodes connected by springs oscillating", "Springs"),
            ("Wave propagation — ripple interference in a 2D grid medium", "Waves"),
            ("Bouncing geometry — shapes reflecting off walls with decay trails", "Reflections"),
            ("N-body gravity — multiple masses attracting each other (pre-computed paths)", "Gravity"),
            ("Diffusion limited aggregation — crystalline branching growth patterns", "Crystal"),
            ("Conway's Game of Life — cellular automata evolution (fully deterministic)", "Automata"),
        ]
    },
    "Cyber / Tech Aesthetics": {
        "description": "Matrix rain, circuit boards, holographic HUDs, neural networks, blockchain",
        "color": ACCENT_BLUE,
        "ideas": [
            ("Matrix digital rain — cascading code columns on black (pure divs, no text rendering)", "Matrix"),
            ("Circuit board traces — PCB-style lines routing and lighting up", "Electronics"),
            ("Holographic HUD overlays — sci-fi interface elements assembling", "Sci-Fi"),
            ("Neural network activation — node layers lighting up with connection pulses", "AI / ML"),
            ("DNA double helix — rotating strand with base-pair animations", "Biology"),
            ("Radar sweep — circular scan with blip detections", "Surveillance"),
            ("Oscilloscope waveforms — audio-visualizer style animated sine stacks", "Audio"),
            ("Binary tree growth — recursive branching data structure building", "Algorithms"),
            ("Glitch / datamosh effects — digital distortion, scanline breaks, pixel displacement", "Glitch Art"),
            ("Blockchain chain formation — linked block nodes assembling in sequence", "Blockchain"),
        ]
    },
    "Space & Cosmos": {
        "description": "Solar systems, supernovae, wormholes, satellite constellations, cosmic web",
        "color": ACCENT_PURPLE,
        "ideas": [
            ("Solar system orrery — planets orbiting the sun to scale with trails", "Solar System"),
            ("Supernova explosion — shockwave rings expanding outward", "Stellar"),
            ("Pulsar beam rotation — lighthouse-style rotating energy beam", "Stellar"),
            ("Comet trails — elongated glowing streaks with particle tails", "Comets"),
            ("Star formation — gas cloud collapsing into stellar ignition", "Stellar"),
            ("Wormhole tunnel — infinite corridor zoom effect", "Sci-Fi"),
            ("Cosmic web — large-scale structure filaments of the universe", "Cosmology"),
            ("Satellite constellation — Starlink-style orbital grid forming", "Satellites"),
            ("Solar flare — magnetic loop eruption on a stylized sun", "Solar"),
            ("Space debris field — orbital debris cloud visualization", "Space"),
        ]
    },
    "Organic / Biological": {
        "description": "Cell division, neural firing, tree growth, bioluminescence, mycelium networks",
        "color": ACCENT_GREEN,
        "ideas": [
            ("Cell division — mitosis animation, single cell splitting into many", "Biology"),
            ("Virus particle — spike protein rotation and binding animation", "Biology"),
            ("Neural firing — action potential propagating down axon branches", "Neuroscience"),
            ("Heartbeat ECG — stylized pulse wave with concentric heart rings", "Medical"),
            ("Tree growth — L-system branching from seed to full canopy", "Nature"),
            ("Root system expansion — underground branching mirror to tree growth", "Nature"),
            ("Coral reef growth — fractal branching colony forming", "Nature"),
            ("Murmuration — starling-flock boid swarm in sweeping formation", "Nature"),
            ("Bioluminescence — deep-sea glow pulses spreading through dark water", "Nature"),
            ("Mycelium network — fungal thread web spreading across dark background", "Nature"),
        ]
    },
    "Architectural / Structural": {
        "description": "Building construction, city skylines, blueprints, structural analysis, urban growth",
        "color": ACCENT_AMBER,
        "ideas": [
            ("Building construction timelapse — floors stacking up, facade filling in", "Construction"),
            ("Bridge cable tension — suspension cable sag and tension animations", "Engineering"),
            ("City skyline reveal — buildings rising from ground level silhouette", "Cityscape"),
            ("Blueprint unfolding — architectural plan lines drawing themselves", "Architecture"),
            ("Structural stress visualization — force flow lines through beam networks", "Engineering"),
            ("Floor plan expansion — rooms adding on as company grows", "Corporate"),
            ("Stadium fill animation — seat-by-seat crowd fill visualization", "Events"),
            ("Urban grid growth — city block grid expanding outward from center", "Urban Planning"),
            ("Tunnel boring — cross-section of tunnel drilling through layers", "Infrastructure"),
            ("Scaffold assembly — modular frame building up from base", "Construction"),
        ]
    },
    "Audio-Visual Style": {
        "description": "Equalizers, waveforms, vinyl records, beat drops — music-driven frame animations",
        "color": ACCENT_PINK,
        "ideas": [
            ("Equalizer bars — frequency spectrum bars pulsing to mathematical curves", "Equalizer"),
            ("Waveform visualizer — oscillating amplitude waveform across screen", "Waveform"),
            ("Vinyl record spin — stylized disc with groove ring animation", "Music"),
            ("Speaker cone pulse — concentric ripple rings from center point", "Audio"),
            ("Beat drop visualization — synchronized geometric explosion at mathematical intervals", "Beat"),
            ("Synthesizer patch — modular signal routing web animating", "Synth"),
            ("Piano roll — note blocks scrolling with key light-up animation", "Music"),
            ("Metronome pendulum — precise rhythmic oscillation with harmonics", "Music"),
        ]
    },
    "Cinematic Transitions": {
        "description": "Liquid morphs, ink drops, lens flares, shockwaves, countdown clocks, neon signs",
        "color": ACCENT_ORANGE,
        "ideas": [
            ("Liquid mercury morph — blob shapes merging and splitting", "Liquid"),
            ("Ink drop diffusion — glow spreading in fluid on dark background", "Fluid"),
            ("Smoke trail dynamics — turbulent particle plume rising", "Smoke"),
            ("Lens flare sweep — anamorphic light streak crossing the frame", "Light"),
            ("Cinematic bars — widescreen letterbox expanding/collapsing with scene", "Cinema"),
            ("Particle logo build — dots converging to form abstract shape", "Logo"),
            ("Shockwave ring pulse — concentric energy rings expanding from center", "Effects"),
            ("Neon sign flicker — tube light charging and illuminating sequence", "Neon"),
            ("Film grain overlay — deterministic noise grain on glowing content", "Film"),
            ("Countdown clock — precision timer with sweeping arc and flash", "Timer"),
        ]
    },
    "Financial / Trading": {
        "description": "Candlesticks, order books, yield curves, portfolio charts, Sankey flows",
        "color": ACCENT_GOLD,
        "ideas": [
            ("Candlestick chart animation — OHLC bars building left to right", "Trading"),
            ("Order book depth — bid/ask walls building and collapsing", "Trading"),
            ("Portfolio donut chart — allocation slices filling and rebalancing", "Portfolio"),
            ("Yield curve shift — 3D surface morphing as rates change", "Fixed Income"),
            ("Bull/bear market cycle — sine-wave market cycle with regime coloring", "Markets"),
            ("Options payoff diagram — payout curve drawing itself", "Derivatives"),
            ("Correlation matrix — heat grid cells filling with color intensity", "Analytics"),
            ("Volatility surface — 3D mesh surface animating as vol changes", "Derivatives"),
            ("Flow of funds Sankey — proportional flow streams between nodes", "Fund Flows"),
            ("IPO roadshow globe — investor city dots lighting in sequence", "Capital Markets"),
        ]
    },
    "Sports & Events": {
        "description": "Shot charts, race tracks, tournament brackets, crowd waves, athlete biometrics",
        "color": ACCENT_RED,
        "ideas": [
            ("Heat map shot chart — basketball/soccer shot location density building", "Analytics"),
            ("Race track position — F1-style lap position tracker", "Racing"),
            ("Marathon runner paths — GPS trace of elite runner routes", "Running"),
            ("Crowd wave — stadium audience wave propagation", "Events"),
            ("Tournament bracket — knockout bracket assembling match by match", "Tournaments"),
            ("Scoreboard ticker — rolling score animation with momentum effects", "Scoring"),
            ("Athlete biometric pulse — heart rate, speed, power output curves", "Biometrics"),
            ("Field formation shift — soccer tactical formation morphing", "Tactics"),
            ("Podium reveal — step-by-step podium rising animation", "Awards"),
            ("Medal tally race — country medal count bar chart race", "Statistics"),
        ]
    },
}

# ═══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── SHEET 1: MASTER INDEX ────────────────────────────────────────────────────

def make_index_sheet(wb):
    ws = wb.create_sheet("📋 Master Index")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT_CYAN

    # Title banner
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "🎬  REMOTION VIDEO IDEAS — MASTER INDEX"
    title_cell.fill = fill(BG_DARK)
    title_cell.font = Font(bold=True, color=ACCENT_CYAN, size=20, name="Calibri")
    title_cell.alignment = center()
    ws.row_dimensions[1].height = 50

    ws.merge_cells("A2:G2")
    sub_cell = ws["A2"]
    sub_cell.value = "Powered by Claude Sonnet 4-6  •  API: https://api.anthropic.com/v1/messages  •  Model: claude-sonnet-4-6"
    sub_cell.fill = fill("1A1A2E")
    sub_cell.font = Font(color=MID_GRAY, size=10, italic=True, name="Calibri")
    sub_cell.alignment = center()
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A3:G3")
    ws["A3"].fill = fill(BG_DARK)
    ws.row_dimensions[3].height = 8

    # Section header — Existing
    ws.merge_cells("A4:G4")
    ws["A4"].value = "EXISTING VIDEO CATEGORIES  (from video_ideas.txt)"
    ws["A4"].fill = fill("162032")
    ws["A4"].font = Font(bold=True, color=ACCENT_CYAN, size=13, name="Calibri")
    ws["A4"].alignment = left()
    ws["A4"].border = Border(bottom=Side(style="medium", color=ACCENT_CYAN))
    ws.row_dimensions[4].height = 30

    # Column headers
    headers = ["#", "Category", "Theme / Focus", "Ideas Count", "Sample Topics", "Status", "Color Code"]
    h_fills = [BG_DARK, BG_DARK, BG_DARK, BG_DARK, BG_DARK, BG_DARK, BG_DARK]
    apply_header_row(ws, 5, headers, h_fills, heights=28)

    row = 6
    existing_list = [
        ("Global Network Animations",   "City/continent connections, supply chains",   len(EXISTING_DATA["Global Network Animations"]["batches"]),    "Connection lines, globe arcs, city nodes",              ACCENT_CYAN),
        ("GPS / Route Tracing",          "Vehicles, logistics, navigation",             len(EXISTING_DATA["GPS / Route Tracing"]["batches"]),           "GPS lines, delivery trucks, shipping lanes",            ACCENT_GREEN),
        ("Data Point Pulses",            "Map markers, sales hotspots",                 len(EXISTING_DATA["Data Point Pulses"]["batches"]),             "Pulsing pins, territory zones, revenue dots",           ACCENT_GOLD),
        ("Globe with Data Overlays",     "Spinning Earth, orbital rings, data arcs",    len(EXISTING_DATA["Globe with Data Overlays"]["batches"]),      "Particle globe, low-poly Earth, data arcs",             ACCENT_BLUE),
        ("Infographic Overlays",         "Charts & bars rising on maps",                len(EXISTING_DATA["Infographic Overlays"]["batches"]),          "Bar charts, pie charts, KPI counters",                  ACCENT_ORANGE),
        ("Flight Paths",                 "Airplane icons, aviation routes",             len(EXISTING_DATA["Flight Paths"]["batches"]),                  "Arc routes, hub-and-spoke, cargo planes",               ACCENT_PINK),
        ("Heatmap Flows",                "Color gradients, density, risk maps",         len(EXISTING_DATA["Heatmap Flows"]["batches"]),                 "Risk zones, revenue density, climate",                  ACCENT_RED),
        ("Connection Webs",              "Hub-and-node, mergers, partnerships",         len(EXISTING_DATA["Connection Webs"]["batches"]),               "Network webs, M&A merges, org charts",                  ACCENT_PURPLE),
        ("Timeline Evolutions",          "Maps changing over time, history",            len(EXISTING_DATA["Timeline Evolutions"]["batches"]),           "Corporate growth, tech adoption, geopolitics",           ACCENT_TEAL),
        ("Icon Swarms",                  "Clustering icons, demographic mapping",       len(EXISTING_DATA["Icon Swarms"]["batches"]),                   "Population swarms, customer flows, workforce",           ACCENT_LIME),
    ]

    for i, (cat, theme, count, sample, color) in enumerate(existing_list, 1):
        bg = ROW_ALT_A if i % 2 == 0 else ROW_ALT_B
        data = [i, cat, theme, count, sample, "✅ In Pipeline", f"#{color}"]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = left() if col > 1 else center()
            if col == 1:
                cell.font = Font(bold=True, color=color, size=11, name="Calibri")
            elif col == 2:
                cell.font = Font(bold=True, color=color, size=11, name="Calibri")
            elif col == 4:
                cell.font = Font(bold=True, color=ACCENT_GOLD, size=12, name="Calibri")
                cell.alignment = center()
            elif col == 6:
                cell.font = Font(color=ACCENT_GREEN, size=10, name="Calibri")
            elif col == 7:
                cell.fill = fill(color)
                cell.font = Font(color=WHITE, size=9, name="Calibri")
                cell.alignment = center()
            else:
                cell.font = Font(color=WHITE, size=10, name="Calibri")
        ws.row_dimensions[row].height = 28
        row += 1

    # Spacer
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].fill = fill(BG_DARK)
    ws.row_dimensions[row].height = 14
    row += 1

    # Section header — New
    ws.merge_cells(f"A{row}:G{row}")
    ws[f"A{row}"].value = "NEW CREATIVE CATEGORIES  (AI-Generated Suggestions)"
    ws[f"A{row}"].fill = fill("1E1628")
    ws[f"A{row}"].font = Font(bold=True, color=ACCENT_GOLD, size=13, name="Calibri")
    ws[f"A{row}"].alignment = left()
    ws[f"A{row}"].border = Border(bottom=Side(style="medium", color=ACCENT_GOLD))
    ws.row_dimensions[row].height = 30
    row += 1

    # Headers again
    apply_header_row(ws, row, headers, h_fills, heights=28)
    row += 1

    new_list = [
        ("Generative Mathematics",  "Fractals, waves, parametric curves",        10, "Mandelbrot, Lissajous, Voronoi, Fibonacci",     ACCENT_INDIGO),
        ("Particle Systems",        "Galaxies, nebulae, fireworks, fluids",       10, "Star trails, aurora, explosion rings, vortex",  ACCENT_CYAN),
        ("Geometric / Abstract",    "Polyhedra, kaleidoscopes, tessellations",    10, "4D hypercube, Penrose tiling, origami",          ACCENT_GOLD),
        ("Physics Simulations",     "Pendulums, orbits, collisions, automata",    10, "Double pendulum, N-body gravity, Game of Life", ACCENT_GREEN),
        ("Cyber / Tech Aesthetics", "Matrix rain, circuits, HUDs, blockchain",    10, "PCB traces, neural nets, glitch effects",       ACCENT_BLUE),
        ("Space & Cosmos",          "Solar systems, supernovae, wormholes",       10, "Orrery, pulsar, cosmic web, satellite grid",    ACCENT_PURPLE),
        ("Organic / Biological",    "Cell division, neurones, tree growth",       10, "Mitosis, mycelium, bioluminescence, murmuration",ACCENT_GREEN),
        ("Architectural / Structural","Buildings, skylines, blueprints, urban",   10, "Construction, bridge, floor plan, grid growth", ACCENT_AMBER),
        ("Audio-Visual Style",      "Equalizers, waveforms, vinyl, beats",         8, "Spectrum bars, piano roll, metronome, synth",  ACCENT_PINK),
        ("Cinematic Transitions",   "Liquid morphs, ink drops, lens flares",      10, "Mercury blob, smoke trail, countdown, neon",    ACCENT_ORANGE),
        ("Financial / Trading",     "Candlesticks, yield curves, Sankey flows",   10, "OHLC chart, order book, vol surface, IPO globe",ACCENT_GOLD),
        ("Sports & Events",         "Shot charts, race tracks, biometrics",       10, "F1 tracker, bracket, formation shift, podium",  ACCENT_RED),
    ]

    for i, (cat, theme, count, sample, color) in enumerate(new_list, 1):
        bg = ROW_ALT_A if i % 2 == 0 else ROW_ALT_B
        data = [i, cat, theme, count, sample, "🆕 New Ideas", f"#{color}"]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = left() if col > 1 else center()
            if col == 1:
                cell.font = Font(bold=True, color=color, size=11, name="Calibri")
            elif col == 2:
                cell.font = Font(bold=True, color=color, size=11, name="Calibri")
            elif col == 4:
                cell.font = Font(bold=True, color=ACCENT_GOLD, size=12, name="Calibri")
                cell.alignment = center()
            elif col == 6:
                cell.font = Font(color=ACCENT_AMBER, size=10, name="Calibri")
            elif col == 7:
                cell.fill = fill(color)
                cell.font = Font(color=WHITE, size=9, name="Calibri")
                cell.alignment = center()
            else:
                cell.font = Font(color=WHITE, size=10, name="Calibri")
        ws.row_dimensions[row].height = 28
        row += 1

    # Total row
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    ws[f"A{row}"].value = f"GRAND TOTAL — 22 Categories"
    ws[f"A{row}"].fill = fill(BG_DARK)
    ws[f"A{row}"].font = Font(bold=True, color=ACCENT_CYAN, size=13, name="Calibri")
    ws[f"A{row}"].alignment = center()
    ws.cell(row=row, column=4).value = 220
    ws.cell(row=row, column=4).fill = fill(BG_DARK)
    ws.cell(row=row, column=4).font = Font(bold=True, color=ACCENT_GOLD, size=14, name="Calibri")
    ws.cell(row=row, column=4).alignment = center()
    ws.merge_cells(f"E{row}:G{row}")
    ws.cell(row=row, column=5).value = "220+ unique video ideas across 22 categories"
    ws.cell(row=row, column=5).fill = fill(BG_DARK)
    ws.cell(row=row, column=5).font = Font(color=MID_GRAY, size=10, italic=True, name="Calibri")
    ws.cell(row=row, column=5).alignment = center()
    ws.row_dimensions[row].height = 36

    set_col_widths(ws, [5, 32, 38, 14, 50, 18, 14])
    ws.freeze_panes = "A6"

# ── SHEET 2: ALL EXISTING IDEAS ──────────────────────────────────────────────

def make_existing_sheet(wb):
    ws = wb.create_sheet("🗂 Existing Ideas")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT_GREEN

    ws.merge_cells("A1:F1")
    ws["A1"].value = "EXISTING VIDEO IDEAS  —  From video_ideas.txt"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = Font(bold=True, color=ACCENT_GREEN, size=18, name="Calibri")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 44

    apply_header_row(ws, 2, ["#", "Category", "Video Idea / Prompt", "Sub-Type", "Status", "Notes"],
                     [BG_DARK]*6, heights=28)

    row = 3
    n = 1
    for cat, data in EXISTING_DATA.items():
        color = CATEGORY_COLORS.get(cat, ACCENT_CYAN)
        for idea, subtype in data["batches"]:
            bg = ROW_ALT_A if n % 2 == 0 else ROW_ALT_B
            vals = [n, cat, idea, subtype, "In Pipeline", ""]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill(bg)
                cell.border = border_thin()
                cell.alignment = left() if col > 1 else center()
                if col == 1:
                    cell.font = Font(bold=True, color=color, size=10, name="Calibri")
                    cell.alignment = center()
                elif col == 2:
                    cell.font = Font(bold=True, color=color, size=10, name="Calibri")
                elif col == 3:
                    cell.font = Font(color=WHITE, size=10, name="Calibri")
                elif col == 4:
                    cell.font = Font(color=ACCENT_GOLD, size=9, italic=True, name="Calibri")
                elif col == 5:
                    cell.font = Font(color=ACCENT_GREEN, size=9, name="Calibri")
                else:
                    cell.font = Font(color=MID_GRAY, size=9, name="Calibri")
            ws.row_dimensions[row].height = 36
            row += 1
            n += 1

    set_col_widths(ws, [5, 32, 75, 20, 14, 20])
    ws.freeze_panes = "A3"

# ── SHEET 3: ALL NEW IDEAS ───────────────────────────────────────────────────

def make_new_sheet(wb):
    ws = wb.create_sheet("✨ New Creative Ideas")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT_GOLD

    ws.merge_cells("A1:F1")
    ws["A1"].value = "NEW CREATIVE VIDEO IDEAS  —  AI-Suggested Categories"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = Font(bold=True, color=ACCENT_GOLD, size=18, name="Calibri")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 44

    apply_header_row(ws, 2, ["#", "Category", "Video Idea / Concept", "Sub-Type", "Complexity", "Priority"],
                     [BG_DARK]*6, heights=28)

    row = 3
    n = 1
    for cat, data in NEW_DATA.items():
        color = data["color"]
        for idea, subtype in data["ideas"]:
            bg = ROW_ALT_A if n % 2 == 0 else ROW_ALT_B
            complexity = "Medium" if "simulation" in idea.lower() or "3D" in idea else "Easy"
            priority = "⭐⭐⭐" if any(k in idea.lower() for k in ["galaxy", "fractal", "neural", "candlestick", "particle"]) else "⭐⭐"
            vals = [n, cat, idea, subtype, complexity, priority]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = fill(bg)
                cell.border = border_thin()
                cell.alignment = left() if col > 1 else center()
                if col == 1:
                    cell.font = Font(bold=True, color=color, size=10, name="Calibri")
                    cell.alignment = center()
                elif col == 2:
                    cell.font = Font(bold=True, color=color, size=10, name="Calibri")
                elif col == 3:
                    cell.font = Font(color=WHITE, size=10, name="Calibri")
                elif col == 4:
                    cell.font = Font(color=ACCENT_GOLD, size=9, italic=True, name="Calibri")
                elif col == 5:
                    cell.font = Font(color=ACCENT_CYAN if val == "Easy" else ACCENT_ORANGE, size=9, name="Calibri")
                else:
                    cell.font = Font(color=ACCENT_GOLD, size=9, name="Calibri")
            ws.row_dimensions[row].height = 42
            row += 1
            n += 1

    set_col_widths(ws, [5, 30, 80, 20, 12, 12])
    ws.freeze_panes = "A3"

# ── SHEET 4: CATEGORY DEEP DIVES ────────────────────────────────────────────

def make_category_sheet(wb, cat_name, description, ideas, color, is_existing=True):
    safe_name = cat_name[:31].replace("/", "-").replace(":", "")
    ws = wb.create_sheet(safe_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = color

    ws.merge_cells("A1:E1")
    ws["A1"].value = cat_name.upper()
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = Font(bold=True, color=color, size=20, name="Calibri")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 48

    ws.merge_cells("A2:E2")
    ws["A2"].value = description
    ws["A2"].fill = fill("151F2E")
    ws["A2"].font = Font(color=MID_GRAY, size=11, italic=True, name="Calibri")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 26

    ws.merge_cells("A3:E3")
    ws["A3"].fill = fill(BG_DARK)
    ws.row_dimensions[3].height = 8

    col_headers = ["#", "Video Idea / Prompt", "Sub-Type", "Status", "Add to Queue?"]
    apply_header_row(ws, 4, col_headers, [BG_DARK]*5, heights=30)

    row = 5
    if is_existing:
        items = ideas  # list of (idea, subtype)
    else:
        items = ideas  # same format

    for i, (idea, subtype) in enumerate(items, 1):
        bg = ROW_ALT_A if i % 2 == 0 else ROW_ALT_B
        status = "In Pipeline" if is_existing else "New Idea"
        vals = [i, idea, subtype, status, "☐"]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            if col == 1:
                cell.alignment = center()
                cell.font = Font(bold=True, color=color, size=12, name="Calibri")
            elif col == 2:
                cell.alignment = left()
                cell.font = Font(color=WHITE, size=10, name="Calibri")
            elif col == 3:
                cell.alignment = center()
                cell.font = Font(color=ACCENT_GOLD, size=9, italic=True, name="Calibri")
            elif col == 4:
                cell.alignment = center()
                fc = ACCENT_GREEN if is_existing else ACCENT_AMBER
                cell.font = Font(color=fc, size=9, name="Calibri")
            else:
                cell.alignment = center()
                cell.font = Font(color=MID_GRAY, size=14, name="Calibri")
        ws.row_dimensions[row].height = 44
        row += 1

    set_col_widths(ws, [5, 82, 20, 14, 16])
    ws.freeze_panes = "A5"

# ── SHEET LAST: PIPELINE TRACKER ────────────────────────────────────────────

def make_tracker_sheet(wb):
    ws = wb.create_sheet("🚀 Pipeline Tracker")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = ACCENT_PURPLE

    ws.merge_cells("A1:H1")
    ws["A1"].value = "🚀  VIDEO PRODUCTION PIPELINE TRACKER"
    ws["A1"].fill = fill(BG_DARK)
    ws["A1"].font = Font(bold=True, color=ACCENT_PURPLE, size=18, name="Calibri")
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 44

    ws.merge_cells("A2:H2")
    ws["A2"].value = "Track each video from idea → generation → render → done"
    ws["A2"].fill = fill("1A1228")
    ws["A2"].font = Font(color=MID_GRAY, size=10, italic=True, name="Calibri")
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 22

    headers = ["#", "Category", "Video Idea", "Component Name", "Status", "API Calls", "Render Time", "Output File"]
    apply_header_row(ws, 3, headers, [BG_DARK]*8, heights=30)

    # Pre-fill with examples
    sample_rows = [
        (1,  "Globe with Data Overlays",   "Spinning Earth with real-time data arcs",               "SpinningEarthArcs",      "✅ DONE",        1, "~45s",   "spinning-earth-arcs.mp4"),
        (2,  "GPS / Route Tracing",         "Glowing neon GPS route line drawing across city grid",  "GlowingNeonGpsRoute",    "❌ FAILED",      1, "—",       "—"),
        (3,  "Flight Paths",                "Airplane arc New York to London",                       "AirplaneArcNewYork",     "⏳ Pending",     0, "—",       "—"),
        (4,  "Generative Mathematics",      "Fibonacci spiral growth animation",                     "FibonacciSpiral",        "🆕 Queue",       0, "—",       "—"),
        (5,  "Particle Systems",            "Galaxy simulation with star trails",                    "GalaxySimulation",       "🆕 Queue",       0, "—",       "—"),
    ]

    for i, row_data in enumerate(sample_rows, 4):
        bg = ROW_ALT_A if i % 2 == 0 else ROW_ALT_B
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill(bg)
            cell.border = border_thin()
            cell.alignment = center() if col in [1, 5, 6, 7] else left()
            if col == 1:
                cell.font = Font(bold=True, color=ACCENT_CYAN, size=11, name="Calibri")
            elif col == 5:
                c = ACCENT_GREEN if "DONE" in str(val) else ACCENT_RED if "FAILED" in str(val) else ACCENT_AMBER
                cell.font = Font(bold=True, color=c, size=10, name="Calibri")
            else:
                cell.font = Font(color=WHITE, size=10, name="Calibri")
        ws.row_dimensions[i].height = 30

    set_col_widths(ws, [5, 30, 50, 30, 14, 12, 14, 30])
    ws.freeze_panes = "A4"

# ── BUILD ALL SHEETS ──────────────────────────────────────────────────────────

make_index_sheet(wb)
make_existing_sheet(wb)
make_new_sheet(wb)

# Individual sheets for each existing category
for cat, data in EXISTING_DATA.items():
    color = CATEGORY_COLORS.get(cat, ACCENT_CYAN)
    make_category_sheet(wb, cat, data["description"], data["batches"], color, is_existing=True)

# Individual sheets for each new category
for cat, data in NEW_DATA.items():
    make_category_sheet(wb, cat, data["description"], data["ideas"], data["color"], is_existing=False)

make_tracker_sheet(wb)

# Save
OUT_PATH = r"D:\remotion_claude_ai\Remotion_Video_Ideas.xlsx"
wb.save(OUT_PATH)
print("Saved: " + OUT_PATH)
print("Sheets: " + str([s.title for s in wb.worksheets]))
